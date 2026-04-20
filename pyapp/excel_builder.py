"""
MergeResult.rows から、納品版と同じ形式の利益計算表Excelを生成する。

列構成:
  A 商品名 / B セット商品コード / C 商品管理番号 / D SS用タイトル
  E 通常価格 / F 原価 / G 送料
  H〜K バリエ内容1〜4 / L 楽天手数料 / M 楽天ペイ利用料 / N ポイント原資 / O NE手数料
  P〜R 10%OFF (価格/利益額/利益率)
  S〜U 20%OFF
  V〜X 30%OFF
  Y〜AA 50%OFF
  AB セット商品コード (B列コピー)
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---- 配色 (納品版と一致) ---------------------------------------------------
COLOR_RED = "EE4D3D"      # セット商品コード/商品管理番号/SS用タイトル
COLOR_GREEN = "3DEE56"    # 通常価格/原価/送料
COLOR_GRAY = "9E9DA7"     # バリエ/手数料系
COLOR_BLUE = "DDEBF7"     # 10% OFF
COLOR_ORANGE = "FCE4D6"   # 20% OFF
COLOR_LIGHTGREEN = "E2F0D9"  # 30% OFF
COLOR_BROWN = "EDE1D1"    # 50% OFF

# ---- 列幅 (納品版と一致) ---------------------------------------------------
COLUMN_WIDTHS = {
    "A": 55.0,
    "B": 35.57,
    "C": 15.14,
    "D": 15.57,
    "E": 10.29,
    "F": 6.0,
    "G": 6.43,
    "H": 26.0,
    "I": 26.0,
    "J": 26.0,
    "K": 26.0,
    "L": 12.71,
    "M": 17.57,
    "N": 15.14,
    "O": 11.14,
    "P": 15.14,
    "Q": 17.57,
    "R": 13.0,
    "S": 15.14,
    "T": 17.57,
    "U": 13.0,
    "V": 15.14,
    "W": 17.57,
    "X": 13.0,
    "Y": 15.14,
    "Z": 17.57,
    "AA": 13.0,
    "AB": 35.57,
}

# ---- ヘッダー定義 -----------------------------------------------------------
HEADERS: list[tuple[str, str, str | None]] = [
    # (列, ラベル, 塗りつぶし色)
    ("A", "商品名", None),
    ("B", "セット商品コード", COLOR_RED),
    ("C", "商品管理番号", COLOR_RED),
    ("D", "SS用タイトル", COLOR_RED),
    ("E", "通常価格", COLOR_GREEN),
    ("F", "原価", COLOR_GREEN),
    ("G", "送料", COLOR_GREEN),
    ("H", "バリエ内容1（送料用）", COLOR_GRAY),
    ("I", "バリエ内容2（送料用）", COLOR_GRAY),
    ("J", "バリエ内容3（送料用）", COLOR_GRAY),
    ("K", "バリエ内容4（送料用）", COLOR_GRAY),
    ("L", "楽天手数料", COLOR_GRAY),
    ("M", "楽天ペイ利用料", COLOR_GRAY),
    ("N", "ポイント原資", COLOR_GRAY),
    ("O", "NE手数料", COLOR_GRAY),
    ("P", "10％OFF価格", COLOR_BLUE),
    ("Q", "10％OFF利益額", COLOR_BLUE),
    ("R", "10％OFF利益率", COLOR_BLUE),
    ("S", "20％OFF価格", COLOR_ORANGE),
    ("T", "20％OFF利益額", COLOR_ORANGE),
    ("U", "20％OFF利益率", COLOR_ORANGE),
    ("V", "30％OFF価格", COLOR_LIGHTGREEN),
    ("W", "30％OFF利益額", COLOR_LIGHTGREEN),
    ("X", "30％OFF利益率", COLOR_LIGHTGREEN),
    ("Y", "50％OFF価格", COLOR_BROWN),
    ("Z", "50％OFF利益額", COLOR_BROWN),
    ("AA", "50％OFF利益率", COLOR_BROWN),
    ("AB", "セット商品コード", None),
]

# 割引列ごとの塗りつぶし色
DISCOUNT_FILLS = {
    "P": COLOR_BLUE, "Q": COLOR_BLUE, "R": COLOR_BLUE,
    "S": COLOR_ORANGE, "T": COLOR_ORANGE, "U": COLOR_ORANGE,
    "V": COLOR_LIGHTGREEN, "W": COLOR_LIGHTGREEN, "X": COLOR_LIGHTGREEN,
    "Y": COLOR_BROWN, "Z": COLOR_BROWN, "AA": COLOR_BROWN,
}

# グループ化する列 (ワンクリック折りたたみ)
GROUP_COLUMNS = ["H", "I", "J", "K", "L", "M", "N", "O"]

# 固定値
FIXED_L = 6        # 楽天手数料
FIXED_M = 3.5      # 楽天ペイ利用料
FIXED_N = 1        # ポイント原資
FIXED_O = 10       # NE手数料

FONT_NAME = "游ゴシック"
FONT_SIZE = 11

THIN = Side(style="thin", color="000000")
BORDER_ALL = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def _fill(color_hex: str) -> PatternFill:
    return PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")


def build_workbook(rows: list[dict[str, Any]]) -> Workbook:
    """処理済みの行リストからWorkbookを構築する。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "利益計算表"

    base_font = Font(name=FONT_NAME, size=FONT_SIZE)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # -------- ヘッダー書き込み --------------------------------------------
    for col_letter, label, color in HEADERS:
        cell = ws[f"{col_letter}1"]
        cell.value = label
        cell.font = base_font
        cell.alignment = center
        cell.border = BORDER_ALL
        if color:
            cell.fill = _fill(color)

    # -------- データ行 ----------------------------------------------------
    excel_row = 2
    for rec in rows:
        if rec.get("_blank"):
            # 空白行: 割引列の背景色だけ保ちつつ値は空のまま
            for col_letter, fill_color in DISCOUNT_FILLS.items():
                cell = ws[f"{col_letter}{excel_row}"]
                cell.fill = _fill(fill_color)
                cell.font = base_font
                cell.border = BORDER_ALL
            excel_row += 1
            continue

        r = excel_row

        # A 商品名
        ws[f"A{r}"] = rec["商品名"]
        ws[f"A{r}"].alignment = left

        # B セット商品コード
        ws[f"B{r}"] = rec["セット商品コード"]
        ws[f"B{r}"].alignment = center

        # C 商品管理番号 (数値化できるなら数値、無理なら文字列)
        mgr = rec["商品管理番号"]
        try:
            ws[f"C{r}"] = int(mgr)
        except (ValueError, TypeError):
            ws[f"C{r}"] = mgr
        ws[f"C{r}"].alignment = center

        # D SS用タイトル (空)
        ws[f"D{r}"] = ""
        ws[f"D{r}"].alignment = left

        # E 通常価格
        ws[f"E{r}"] = rec["通常価格"]
        ws[f"E{r}"].number_format = "0"
        ws[f"E{r}"].alignment = center

        # F 原価
        if rec["原価"] is not None:
            ws[f"F{r}"] = rec["原価"]
        ws[f"F{r}"].number_format = "0"
        ws[f"F{r}"].alignment = center

        # G 送料 (空欄だが通貨書式)
        ws[f"G{r}"].number_format = r'\¥#,##0'
        ws[f"G{r}"].alignment = center

        # H〜K バリエ
        ws[f"H{r}"] = rec["バリエ1"]
        ws[f"I{r}"] = rec["バリエ2"]
        ws[f"J{r}"] = rec["バリエ3"]
        ws[f"K{r}"] = rec["バリエ4"]
        for c in ["H", "I", "J", "K"]:
            ws[f"{c}{r}"].alignment = center

        # L〜O 固定値
        ws[f"L{r}"] = FIXED_L
        ws[f"L{r}"].number_format = "0"
        ws[f"M{r}"] = FIXED_M
        ws[f"M{r}"].number_format = "0.0"
        ws[f"N{r}"] = FIXED_N
        ws[f"N{r}"].number_format = "0"
        ws[f"O{r}"] = FIXED_O
        ws[f"O{r}"].number_format = "0"
        for c in ["L", "M", "N", "O"]:
            ws[f"{c}{r}"].alignment = center

        # P-AA 割引系の数式
        # 10% OFF
        ws[f"P{r}"] = f"=ROUNDDOWN(E{r}*0.9,-1)"
        ws[f"Q{r}"] = f"=ROUNDDOWN(P{r}-F{r}-G{r}-(P{r}*0.01*L{r})-(P{r}*0.01*M{r})-(P{r}*0.01*N{r})-O{r},0)"
        ws[f"R{r}"] = f'=IFERROR(ROUNDDOWN((Q{r}/P{r})*100,0),"")'
        # 20% OFF
        ws[f"S{r}"] = f"=ROUNDDOWN(E{r}*0.8,-1)"
        ws[f"T{r}"] = f"=ROUNDDOWN(S{r}-F{r}-G{r}-(S{r}*0.01*L{r})-(S{r}*0.01*M{r})-(S{r}*0.01*N{r})-O{r},0)"
        ws[f"U{r}"] = f'=IFERROR(ROUNDDOWN((T{r}/S{r})*100,0),"")'
        # 30% OFF
        ws[f"V{r}"] = f"=ROUNDDOWN(E{r}*0.7,-1)"
        ws[f"W{r}"] = f"=ROUNDDOWN(V{r}-F{r}-G{r}-(V{r}*0.01*L{r})-(V{r}*0.01*M{r})-(V{r}*0.01*N{r})-O{r},0)"
        ws[f"X{r}"] = f'=IFERROR(ROUNDDOWN((W{r}/V{r})*100,0),"")'
        # 50% OFF
        ws[f"Y{r}"] = f"=ROUNDDOWN(E{r}*0.5,-1)"
        ws[f"Z{r}"] = f"=ROUNDDOWN(Y{r}-F{r}-G{r}-(Y{r}*0.01*L{r})-(Y{r}*0.01*M{r})-(Y{r}*0.01*N{r})-O{r},0)"
        ws[f"AA{r}"] = f'=IFERROR(ROUNDDOWN((Z{r}/Y{r})*100,0),"")'

        # 価格セルは太字、すべて中央揃え、整数書式、塗りつぶし
        for col in ["P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA"]:
            cell = ws[f"{col}{r}"]
            cell.alignment = center
            cell.number_format = "0"
            cell.fill = _fill(DISCOUNT_FILLS[col])
        # 価格セル(P/S/V/Y)のみ太字
        for col in ["P", "S", "V", "Y"]:
            ws[f"{col}{r}"].font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)

        # AB セット商品コード (B列のコピー)
        ws[f"AB{r}"] = rec["セット商品コード"]
        ws[f"AB{r}"].alignment = center

        # 全セルにフォント・罫線・中央揃え(左揃え指定のA,Dは上書き済み)
        for col_idx in range(1, 29):
            col_letter = get_column_letter(col_idx)
            cell = ws[f"{col_letter}{r}"]
            if cell.font.name != FONT_NAME or cell.font.bold:
                # 太字指定済みはそのまま
                pass
            else:
                cell.font = base_font
            cell.border = BORDER_ALL

        excel_row += 1

    # -------- 列幅 -------------------------------------------------------
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # -------- グループ化 (H〜O) ------------------------------------------
    # summary_right=False: グループのトグルボタンをグループの左側に置く
    # (納品版は summary_right=True だが、これはグループのトグル位置の話)
    for col_letter in GROUP_COLUMNS:
        ws.column_dimensions[col_letter].outline_level = 1
        ws.column_dimensions[col_letter].hidden = False  # 最初は開いた状態

    # -------- ヘッダー行にフィルタ（任意） ---------------------------------
    # ws.auto_filter.ref = f"A1:AB{excel_row-1}"  # ここは好みで

    # -------- フリーズペイン (ヘッダー固定) --------------------------------
    ws.freeze_panes = "A2"

    return wb


def build_xlsx_bytes(rows: list[dict[str, Any]]) -> bytes:
    """Workbookをバイト列として返す。"""
    wb = build_workbook(rows)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
