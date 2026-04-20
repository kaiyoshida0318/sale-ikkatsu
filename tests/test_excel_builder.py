"""excel_builder の単体テスト。"""
import io
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "pyapp"))

from excel_builder import build_xlsx_bytes  # noqa: E402
from merger import merge_all  # noqa: E402


SAMPLE_DIR = ROOT / "sample_data"


def _read(name: str) -> bytes:
    return (SAMPLE_DIR / name).read_bytes()


def test_build_produces_valid_xlsx():
    res = merge_all(
        rms_bytes=_read("RMSデータ.csv"),
        set_item_bytes=_read("NE-セット商品コード.csv"),
        set_component_bytes=_read("NE-セット構成物コード.csv"),
        non_set_bytes=_read("NE-セット無しコード.csv"),
    )
    xlsx = build_xlsx_bytes(res.rows)
    assert xlsx.startswith(b"PK")

    wb = load_workbook(io.BytesIO(xlsx))
    ws = wb.active
    assert ws.title == "利益計算表"
    assert ws["A1"].value == "商品名"
    assert ws["P2"].value == "=ROUNDDOWN(E2*0.9,-1)"
